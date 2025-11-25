#!/usr/bin/env python3
"""
Batch command generator/executor for SAGE experiments.

Given an Excel sheet that lists (model_name, layer, feature_index), this script:
1. Filters rows matching the target model/layer.
2. Randomly samples N feature indices (default 10, deterministic seed).
3. For each sampled feature and each requested top_k value, builds a command that
   runs `main.py` with the appropriate parameters, cycling across available CUDA
   devices.
4. Prints the full schedule for validation, writes it to a markdown log, and
   optionally executes the commands at a fixed interval (default 20 minutes),
   logging progress as it goes.
"""

from __future__ import annotations

import argparse
import datetime as dt
import os
import random
import subprocess
import sys
import time
from dataclasses import dataclass
from itertools import cycle
from pathlib import Path
from typing import Iterable, List, Sequence
import re


@dataclass(frozen=True)
class FeatureRow:
    model_name: str
    layer: str
    feature_index: int


@dataclass
class CommandPlan:
    idx: int
    feature: int
    top_k: int
    cuda: int
    planned_time: dt.datetime
    command: str


def read_feature_rows(xlsx_path: Path) -> List[FeatureRow]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency should exist
        raise SystemExit(
            "openpyxl is required to read the feature spreadsheet. "
            "Install it with `pip install openpyxl`."
        ) from exc
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    header_map = {name: idx for idx, name in enumerate(headers)}
    required = ["model_name", "layer", "feature_index"]
    missing = [h for h in required if h not in header_map]
    if missing:
        raise ValueError(f"Missing columns in Excel file: {', '.join(missing)}")

    rows: List[FeatureRow] = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        if all(value is None for value in excel_row):
            continue
        try:
            rows.append(
                FeatureRow(
                    model_name=str(excel_row[header_map["model_name"]]).strip(),
                    layer=str(excel_row[header_map["layer"]]).strip(),
                    feature_index=int(excel_row[header_map["feature_index"]]),
                )
            )
        except (TypeError, ValueError):
            continue  # skip malformed rows

    if not rows:
        raise ValueError("No valid rows found in Excel file.")
    return rows


def sample_features(
    rows: Sequence[FeatureRow],
    target_model: str,
    target_layer: str,
    sample_size: int,
    seed: int,
) -> List[int]:
    matching = [
        row.feature_index
        for row in rows
        if row.model_name == target_model and row.layer == target_layer
    ]
    if len(matching) < sample_size:
        raise ValueError(
            f"Requested {sample_size} features but only {len(matching)} available for "
            f"model={target_model}, layer={target_layer}."
        )
    random.seed(seed)
    return random.sample(matching, sample_size)


def extract_layer_index(layer_descriptor: str) -> int:
    match = re.search(r"\d+", layer_descriptor)
    if not match:
        raise ValueError(f"Could not extract numeric layer index from '{layer_descriptor}'")
    return int(match.group())


def collect_features_from_results(
    root_dir: Path,
    agent_llm: str,
    target_llm: str,
    layer_descriptor: str,
) -> List[int]:
    layer_index = extract_layer_index(layer_descriptor)
    target_dir = (
        root_dir
        / agent_llm
        / target_llm.replace("/", "_")
        / f"layer_{layer_index}"
    )

    if not target_dir.exists():
        return []

    features: List[int] = []
    for child in target_dir.iterdir():
        if child.is_dir() and child.name.startswith("feature_"):
            try:
                features.append(int(child.name.split("_", 1)[1]))
            except (ValueError, IndexError):
                continue
    return sorted(features)


def sample_from_list(items: Sequence[int], sample_size: int, seed: int) -> List[int]:
    if sample_size > len(items):
        raise ValueError(f"Requested {sample_size} items but only {len(items)} available.")
    random.seed(seed)
    return random.sample(list(items), sample_size)


def build_command(
    working_dir: Path,
    cuda: int,
    target_llm: str,
    neuronpedia_model_id: str,
    neuronpedia_source: str,
    sae_path: str,
    feature: int,
    use_api: bool,
    agent_llm: str,
    max_rounds: int,
    timeout_minutes: int,
    top_k: int,
    device: str,
    path2save: str,
    debug: bool,
) -> str:
    args = [
        f"CUDA_VISIBLE_DEVICES={cuda}",
        "python3",
        "main.py",
        "--target_llm",
        target_llm,
        "--neuronpedia_model_id",
        neuronpedia_model_id,
        "--neuronpedia_source",
        neuronpedia_source,
        "--sae_path",
        sae_path,
        "--features",
        f"layer7={feature}",
        "--use_api_for_activations",
        str(use_api).lower(),
        "--agent_llm",
        agent_llm,
        "--max_rounds",
        str(max_rounds),
        "--timeout_minutes",
        str(timeout_minutes),
        "--device",
        device,
        "--top_k",
        str(top_k),
        "--path2save",
        path2save,
    ]
    if debug:
        args.append("--debug")
    return f"cd {working_dir} && " + " ".join(args)


def create_schedule(
    features: Sequence[int],
    top_k_values: Sequence[int],
    cuda_pool: Iterable[int],
    start_time: dt.datetime,
    interval_minutes: int,
    device_template: str,
    path_default: str,
    path_top5: str,
    **command_kwargs,
) -> List[CommandPlan]:
    schedule: List[CommandPlan] = []
    cuda_cycle = cycle(cuda_pool)
    idx = 1
    for feature in features:
        for top_k in top_k_values:
            cuda_id = next(cuda_cycle)
            planned_time = start_time + dt.timedelta(minutes=interval_minutes * (idx - 1))
            device_value = device_template.format(cuda=cuda_id)
            path2save = path_top5 if top_k == 5 else path_default
            cmd = build_command(
                feature=feature,
                top_k=top_k,
                cuda=cuda_id,
                device=device_value,
                path2save=path2save,
                **command_kwargs,
            )
            schedule.append(
                CommandPlan(
                    idx=idx,
                    feature=feature,
                    top_k=top_k,
                    cuda=cuda_id,
                    planned_time=planned_time,
                    command=cmd,
                )
            )
            idx += 1
    return schedule


def print_schedule(schedule: Sequence[CommandPlan]) -> None:
    for plan in schedule:
        ts = plan.planned_time.strftime("%Y-%m-%d %H:%M")
        print(
            f"[{plan.idx:02d}] {ts} | feature {plan.feature} | top_k={plan.top_k} | cuda:{plan.cuda}"
        )
        print(plan.command)
        print()


def write_markdown(schedule: Sequence[CommandPlan], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# SAGE Batch Command Schedule",
        "",
        "| # | Feature | top_k | CUDA | Planned Start | Command |",
        "|---|---------|-------|------|---------------|---------|",
    ]
    for plan in schedule:
        cmd_short = plan.command.split("&&", 1)[-1].strip()
        lines.append(
            f"| {plan.idx} | {plan.feature} | {plan.top_k} | {plan.cuda} | "
            f"{plan.planned_time.strftime('%Y-%m-%d %H:%M')} | `{cmd_short}` |"
        )
    output_path.write_text("\n".join(lines), encoding="utf-8")


def execute_schedule(
    schedule: Sequence[CommandPlan],
    log_path: Path,
    interval_minutes: int,
) -> None:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    for idx, plan in enumerate(schedule):
        now = dt.datetime.now()
        if now < plan.planned_time:
            wait_seconds = (plan.planned_time - now).total_seconds()
            time.sleep(wait_seconds)
        start_ts = dt.datetime.now()
        result = subprocess.run(plan.command, shell=True)
        end_ts = dt.datetime.now()
        with log_path.open("a", encoding="utf-8") as log_file:
            log_file.write(
                f"[{plan.idx:02d}] feature={plan.feature} top_k={plan.top_k} cuda={plan.cuda} "
                f"start={start_ts.isoformat()} end={end_ts.isoformat()} "
                f"returncode={result.returncode}\n"
            )
        if idx < len(schedule) - 1:
            time.sleep(interval_minutes * 60)


def parse_args(argv: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate/execute SAGE batch commands.")
    parser.add_argument("--xlsx", type=Path, default=Path("topk.xlsx"), help="Path to Excel file.")
    parser.add_argument("--target-llm", default="qwen3-4b", help="Target LLM / model name.")
    parser.add_argument("--layer", default="7-transcoder-hp", help="Layer identifier in Excel.")
    parser.add_argument("--sample-size", type=int, default=10, help="Number of features to sample.")
    parser.add_argument("--seed", type=int, default=20241117, help="Deterministic sampling seed.")
    parser.add_argument(
        "--index-source",
        choices=["results", "excel"],
        default="results",
        help="Where to read candidate feature indices from.",
    )
    parser.add_argument(
        "--use-all-features",
        action="store_true",
        help="Use all discovered features (no sampling). Only applies when --index-source=results.",
    )
    parser.add_argument(
        "--top-k-values",
        type=int,
        nargs="+",
        default=[5, 15],
        help="List of top_k values to run per feature.",
    )
    parser.add_argument(
        "--gpus",
        type=int,
        nargs="+",
        default=list(range(8)),
        help="CUDA device IDs to cycle through.",
    )
    parser.add_argument("--interval-minutes", type=int, default=5, help="Launch interval.")
    parser.add_argument(
        "--working-dir",
        type=Path,
        default=Path("/common/users/wx139/code/SAGE/sga"),
        help="Directory containing main.py.",
    )
    parser.add_argument("--agent-llm", default="gpt-5", help="Agent LLM name.")
    parser.add_argument("--max-rounds", type=int, default=14, help="Max rounds for main.py.")
    parser.add_argument("--timeout-minutes", type=int, default=30, help="Timeout per run.")
    parser.add_argument("--use-api", action="store_true", default=True, help="Use API flag.")
    parser.add_argument("--device-template", default="cuda:{cuda}", help="--device value template.")
    parser.add_argument("--sae-path", default="7-transcoder-hp", help="SAE path/identifier.")
    parser.add_argument("--path-default", default="results", help="Path passed to --path2save for top_k != 5 runs.")
    parser.add_argument("--path-top5", default="topk_result", help="Path passed to --path2save when top_k == 5.")
    parser.add_argument(
        "--results-source-path",
        type=Path,
        default=Path("results"),
        help="Root directory containing existing results (used when --index-source=results).",
    )
    parser.add_argument("--schedule-output", type=Path, default=Path("command_schedule.md"))
    parser.add_argument("--log-file", type=Path, default=Path("batch_run.log"))
    parser.add_argument("--execute", action="store_true", help="Execute commands after generation.")
    parser.add_argument("--debug", action="store_true", help="Pass --debug to main.py.")
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])

    if args.index_source == "excel":
        rows = read_feature_rows(args.xlsx)
        sampled = sample_features(
            rows,
            target_model=args.target_llm,
            target_layer=args.layer,
            sample_size=args.sample_size,
            seed=args.seed,
        )
    else:
        candidates = collect_features_from_results(
            root_dir=args.results_source_path,
            agent_llm=args.agent_llm,
            target_llm=args.target_llm,
            layer_descriptor=args.layer,
        )
        if not candidates:
            raise SystemExit(
                f"No existing features found under {args.results_source_path} "
                f"for agent={args.agent_llm}, model={args.target_llm}, layer={args.layer}"
            )
        if args.use_all_features:
            sampled = list(candidates)
        else:
            sampled = sample_from_list(candidates, args.sample_size, args.seed)
    command_kwargs = dict(
        working_dir=args.working_dir,
        target_llm=args.target_llm,
        neuronpedia_model_id=args.target_llm,
        neuronpedia_source=args.layer,
        sae_path=args.sae_path,
        use_api=args.use_api,
        agent_llm=args.agent_llm,
        max_rounds=args.max_rounds,
        timeout_minutes=args.timeout_minutes,
        debug=args.debug,
    )
    schedule = create_schedule(
        features=sampled,
        top_k_values=args.top_k_values,
        cuda_pool=args.gpus,
        start_time=dt.datetime.now(),
        interval_minutes=args.interval_minutes,
        device_template=args.device_template,
        path_default=str(args.path_default),
        path_top5=str(args.path_top5),
        **command_kwargs,
    )

    print_schedule(schedule)
    write_markdown(schedule, args.schedule_output)

    if args.execute:
        execute_schedule(schedule, args.log_file, args.interval_minutes)
    return 0


if __name__ == "__main__":  # pragma: no cover
    sys.exit(main())

